import PyPDF2
import re
import openpyxl

def read_specific_page(file_path, page_number):
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            if page_number < 1 or page_number > len(pdf_reader.pages):
                return f"Page {page_number} does not exist in the PDF."
            page = pdf_reader.pages[page_number - 1]
            text = page.extract_text()
            return text
    except FileNotFoundError:
        return f"File {file_path} not found."
    except Exception as e:
        return str(e)


def extract_savings_info(text):
    savings_potential_pattern = r"Savings potential\s+(\d{1,2}\.\d+%)"
    mt_c02_savings_pattern = r"([\d,]+)\s*MT Fuel savings"
    mt_fuel_savings_pattern = rf"(\d+)\s*\n\s*{re.escape('Main Engine')}"
    roi_pattern = r"Return of investment\s+\(years\)\s+(\d+\.\d+)"
    savings_potential = re.search(savings_potential_pattern, text)
    mt_c02_savings = re.search(mt_c02_savings_pattern, text)
    mt_fuel_savings = re.search(mt_fuel_savings_pattern, text)
    roi = re.search(roi_pattern, text)
    return {
        "savings_potential": savings_potential.group(1) if savings_potential else None,
        "mt_co2_savings": mt_c02_savings.group(1).replace(',', '') if mt_c02_savings else None,
        "roi": roi.group(1) if roi else None,
        "mt_fuel_savings":mt_fuel_savings.group(1).replace(',', '') if mt_fuel_savings else None
    }

def append_data_excel(data):
    try:
        # Load workbook, create if it doesn't exist
        try:
            book = openpyxl.load_workbook('data.xlsx')
        except FileNotFoundError:
            book = openpyxl.Workbook()
        
        active = book.active
        
        # Append headers if the sheet is empty
        if active.max_row == 1:
            active.append(['Savings Potential', 'MT CO2 Savings', 'MT Fuel Savings', 'ROI'])
        
        # Append the data
        active.append([
            data.get('savings_potential'),
            data.get('mt_co2_savings'),
            data.get('mt_fuel_savings'),
            data.get('roi')
        ])
        
        # Save the workbook
        book.save('data.xlsx')
    except Exception as e:
        print(f"An error occurred while appending data to Excel: {e}")

if __name__ == '__main__':
    file_path = 'sample.pdf'  # Replace with your PDF file path
    page_number = 8  # Specify the page number you want to read
    text = read_specific_page(file_path, page_number)
    data = extract_savings_info(text)
    append_data_excel(data)
